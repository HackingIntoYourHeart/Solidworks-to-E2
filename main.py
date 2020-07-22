if True:
	import os
	import urllib
	import urllib2
	from sys import argv

	try:
		import openpyxl
	except:
		print("openpyxl NOT INSTALLED. Installing...")
		os.system("C:\Python27\Scripts\pip.exe install openpyxl")
		try:
			import openpyxl
		except:
			print("FAILED to install openpyxl... Trying alternate method...")
			PYTHON = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Python27')
			os.system(PYTHON + "\\Scripts\\pip.exe install openpyxl")
			try:
				import openpyxl
			except:
				print("FAILED to install openpyxl... Closing...")
				raw_input()
				exit()
				
	try:
		import win32com.client as win32
	except:
		print("pywin32 NOT INSTALLED. Installing...")
		os.system("C:\Python27\Scripts\pip.exe install pywin32")
		try:
			import win32com.client as win32
		except:
			print("FAILED to install pywin32... Trying alternate method...")
			PYTHON = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Python27')
			os.system(PYTHON + "\\Scripts\\pip.exe install pywin32")
			try:
				import win32com.client as win32
			except:
				print("FAILED to install pywin32... Trying ANOTHER alternate method...")
				os.system("C:\Python27\Scripts\pip.exe install pypiwin32")
				try:
					import win32com.client as win32
				except:
					print("FAILED to install pywin32... Closing...")
					raw_input()
					exit()

	from openpyxl import Workbook
	from openpyxl import load_workbook
	from openpyxl.workbook import Workbook
	from openpyxl.styles import Color, PatternFill, Font, Border, Fill, colors, Alignment
	from openpyxl.cell import Cell
	from openpyxl.styles.borders import Border, Side
	
	desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')

	resource = str(argv[1])
	inputName = (resource.split("\\")[-1]).split(".")[0]
	
	############################################################################
	if resource[-1] != "x":
		print "Converting OLD xls to NEW xlsx..."
		excel = win32.gencache.EnsureDispatch('Excel.Application')
		wbold = excel.Workbooks.Open(resource)

		wbold.SaveAs(resource+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
		wbold.Close()                               #FileFormat = 56 is for .xls extension
		excel.Application.Quit()
		resource = resource + "x"
		print "Conversion COMPLETE!"
	############################################################################
	
	resource2 = os.path.dirname(os.path.realpath(__file__)) + "\\rc\\rcBOM.xlsx"

	# Source Excel Doc. for conversion
	wb = load_workbook(resource)
	ws = wb.worksheets[0]

	# BOM for E2 Template
	wb2 = load_workbook(resource2)
	
	#-------------------------#
	outPART = wb2["PART"]
	outBOM = wb2["BOM"]
	outROUTING = wb2["ROUTING"]
	outOUTSIDE = wb2["OUTSIDE"]
	outMISC = wb2["MISC"]
	#-------------------------#
	
	inBOM = ws
	
	SAVE = wb2
	
	'''# SEARCH
	wb3 = load_workbook(filename)
	ws3 = wb3.worksheets[0]'''
	
	def clean(string):
		try:
			#print(string)
			string = str(string)
			string = string.lower()
			string = string.strip()
			return string
		except:
			return "Error"
	
	'''def search(q):
		for i in range(2,ws3.max_row):
			if clean(q) == clean(ws3.cell(row=i, column=2).value):
				return i
			else:
				return 0'''
				
	def save(title):
		SAVE.save(desktop + '\\' + title + ".xlsx")
		
	###############################################
	
	####
	NUM_L = "D"
	DESC_L = "E"
	QTY_L = "H"
	####
	
	for i in range(2,inBOM.max_row):
		NUM = (inBOM.cell(row=i, column=1).value)
		DESC = (inBOM.cell(row=i, column=2).value)
		QTY = (inBOM.cell(row=i, column=3).value)
		
		print "Writing NUMBER:", NUM, "to sheet."
		outBOM[NUM_L + str(i)] = NUM
		print "Writing DESCRIPTION:", DESC, "to sheet."
		outBOM[DESC_L + str(i)] = DESC
		print "Writing QUANTITY:", QTY, "to sheet."
		outBOM[QTY_L + str(i)] = QTY

		outBOM["B" + str(i)] = str(inputName.split(" ")[0]) + " " + str(inputName.split(" ")[1])
		outBOM["C" + str(i)] = "DEFAULT"
		outBOM["F" + str(i)] = "DEFAULT"
		outBOM["G" + str(i)] = "DEFAULT"
		outBOM["A" + str(i)] = str(i - 1)
	print "CONVERSION PROCESS DONE."
	
	###############################################
	
	save(inputName)
	print "FILE SAVED SUCCESSFULLY! BOM CONVERTED."
	raw_input()
